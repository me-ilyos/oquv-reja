import math
import re
from pathlib import Path

import openpyxl

from parser.models import (
    Alternative,
    ColumnLayout,
    Course,
    ParseError,
    ParseResult,
    SelectiveSlot,
)

COURSE_NUM_RE = re.compile(r"^\d+\.\d+(\.\d+)?$")
SECTION_MARKER_RE = re.compile(r"^(\d+)(?:\.0+)?\.?$")
CODE_RE = re.compile(r"\d{6,9}")
SEP_AFTER_CODE_RE = re.compile(r"^[\s:\-–—]+")
DURATION_RE = re.compile(r"(\d+(?:\.\d+)?)\s*yil")
START_YEAR_RE = re.compile(r"(\d{4})\s*[-/–]\s*\d{4}")
APOSTROPHE_RE = re.compile(r"['\u2018\u2019\u02bb`]")
# Ministry staff occasionally type a header on a Russian keyboard layout,
# leaving a visually-identical Cyrillic letter where a Latin one was meant
# (e.g. "\u041areditlar" with a Cyrillic \u041a) \u2014 invisible to the eye, fatal to a
# substring search. Map the common look-alikes to their Latin equivalent.
CYRILLIC_LOOKALIKES = str.maketrans(
    {
        "\u0410": "A",
        "\u0412": "B",
        "\u0415": "E",
        "\u041a": "K",
        "\u041c": "M",
        "\u041d": "H",
        "\u041e": "O",
        "\u0420": "P",
        "\u0421": "C",
        "\u0422": "T",
        "\u0423": "Y",
        "\u0425": "X",
        "\u0430": "a",
        "\u0435": "e",
        "\u043e": "o",
        "\u0440": "p",
        "\u0441": "c",
        "\u0443": "y",
        "\u0445": "x",
    }
)

EDU_TYPE_CANONICAL = {
    "kunduzgi": "Kunduzgi",
    "kechki": "Kechki",
    "kechgi": "Kechki",
    "sirtqi": "Sirtqi",
    "sirtki": "Sirtqi",
}

WEEKLY_GROUP_LABELS = ("haftalar soni", "kurslardagi haftalar")
CREDIT_GROUP_LABELS = ("kredit",)
SKIP_ROW_LABELS = (
    "malakaviy amaliyot",
    "malaka amaliyot",
    "yakuniy davlat attestatsiyasi",
    "hammasi",
)
BREAKDOWN_FIELDS = ("classroom", "lecture", "practice", "lab", "seminar", "course_proj")

_UNPARSEABLE = object()


def is_course_number(value) -> bool:
    if value is None:
        return False
    return bool(COURSE_NUM_RE.match(str(value).strip()))


def section_prefix(course_num) -> str:
    return str(course_num).strip().split(".")[0]


def _normalize_text(value: str) -> str:
    """Casefold, collapse the apostrophe family to a single character, and
    map stray Cyrillic look-alike letters to Latin, so matching is stable
    across the glyph variants and keyboard-layout slips seen in `sources/`."""
    value = APOSTROPHE_RE.sub("'", value).translate(CYRILLIC_LOOKALIKES)
    return value.casefold()


def _normalize_edu_type(raw: str) -> str | None:
    """Match a ta'lim shakli value to its canonical form (Kunduzgi/Kechki/
    Sirtqi) by substring containment, so trailing text like '(4 yil)' doesn't
    block the match. Returns None when nothing matches."""
    norm = _normalize_text(raw).strip()
    for needle, canonical in EDU_TYPE_CANONICAL.items():
        if needle in norm:
            return canonical
    return None


def _normalize_section_marker(value) -> str | None:
    """Normalize a section-marker cell ('1.00', '1,0', '1.00.', ...) to its
    bare section digit string, or None if the cell isn't a marker at all."""
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    match = SECTION_MARKER_RE.match(text)
    return match.group(1) if match else None


def find_cell_containing(ws, substring: str):
    needle = _normalize_text(substring)
    for row in ws.iter_rows(max_row=15):
        for cell in row:
            if isinstance(cell.value, str) and needle in _normalize_text(cell.value):
                return cell
    return None


def value_after_dash(raw: str) -> str:
    parts = re.split(r"\s*[-–—:]\s*", raw, maxsplit=1)
    return parts[1].strip() if len(parts) > 1 else raw.strip()


def _extract_code_name(raw: str) -> tuple[str, str] | None:
    """Find a 6-9 digit direction code in `raw` and the name text after it."""
    match = CODE_RE.search(raw)
    if match is None:
        return None
    code = match.group()
    name = SEP_AFTER_CODE_RE.sub("", raw[match.end() :]).strip()
    return code, name


def resolve_direction(ws, warnings: list[str]) -> tuple[str, str] | None:
    """Find and parse the program-direction cell into (code, name).

    Returns None when no direction cell is present at all (caller falls back
    to the filename). The code/name pair may appear in the label cell itself,
    the cell to its right, or the cell below it — try each in turn and use
    the first that yields a plausible code.
    """
    cell = find_cell_containing(ws, "nalishi")
    if cell is None:
        return None
    candidates = [
        cell.value,
        ws.cell(row=cell.row, column=cell.column + 1).value,
        ws.cell(row=cell.row + 1, column=cell.column).value,
    ]
    for raw in candidates:
        if not isinstance(raw, str):
            continue
        extracted = _extract_code_name(raw)
        if extracted is not None:
            return extracted
    warnings.append(f"yo'nalishi kodi topilmadi: {cell.value!r}")
    return "", value_after_dash(str(cell.value or ""))


def to_pascal_case(name: str) -> str:
    return "".join(re.sub(r"[^a-zA-Z0-9]", "", w).capitalize() for w in name.split())


def extract_start_year(raw: str) -> str:
    match = START_YEAR_RE.search(raw)
    return match.group(1) if match else ""


def read_metadata(ws, warnings: list[str]) -> tuple[str, str, str]:
    """Return (degree, duration, edu_type) from the metadata region, each the
    text after the label's separator or '' when its cell is absent."""
    degree_cell = find_cell_containing(ws, "Akademik daraja")
    duration_cell = find_cell_containing(ws, "muddati")
    edu_type_cell = find_cell_containing(ws, "ta'lim shakli")
    degree = value_after_dash(degree_cell.value) if degree_cell else ""
    duration = value_after_dash(duration_cell.value) if duration_cell else ""
    if not degree:
        warnings.append("akademik daraja topilmadi")
    if edu_type_cell is None:
        warnings.append("ta'lim shakli topilmadi")
        return degree, duration, ""
    raw_edu_type = value_after_dash(edu_type_cell.value)
    edu_type = _normalize_edu_type(raw_edu_type)
    if edu_type is None:
        warnings.append(f"ta'lim shakli tanilmadi: {raw_edu_type!r}")
        return degree, duration, ""
    return degree, duration, edu_type


def find_header_col(ws, keyword: str, max_row: int, min_row: int = 1) -> int | None:
    """Return 0-based column index of first header cell containing keyword, or None."""
    needle = _normalize_text(keyword)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        for cell in row:
            if isinstance(cell.value, str) and needle in _normalize_text(cell.value):
                return cell.column - 1
    return None


def extract_program_years(ws) -> float:
    """Return program duration in years (e.g. 3, or 4.5 for a 9-semester
    program). Structural: the credit-column window is derived from it."""
    cell = find_cell_containing(ws, "muddati")
    if cell is None:
        raise ParseError("o'qish muddati topilmadi")
    text = str(cell.value or "").replace(",", ".")
    match = DURATION_RE.search(text)
    if match is None:
        raise ParseError(f"o'qish muddati aniqlanmadi: {cell.value!r}")
    return float(match.group(1))


def _find_anchor(ws, warnings: list[str]) -> tuple[int, int]:
    """Locate the '1.00' section-marker cell. Returns (0-based num_col,
    1-based row). Raises ParseError when no candidate is found; when several
    numeric '1'-like cells exist (e.g. inside the academic-calendar zone),
    prefers the one whose row also contains 'majburiy'."""
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            if _normalize_section_marker(cell.value) == "1":
                matches.append((cell.column - 1, cell.row))
    if not matches:
        raise ParseError("'1.00' bo'lim belgisi topilmadi")

    for col, row_num in matches:
        row_values = next(
            ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True)
        )
        if any(
            isinstance(v, str) and "majburiy" in _normalize_text(v) for v in row_values
        ):
            return col, row_num
    warnings.append(
        "'1.00' bo'lim belgisi 'majburiy' yorlig'isiz topildi;"
        " birinchi mos katakka tayanildi"
    )
    return matches[0]


def _to_float(value) -> float | None:
    """Coerce an Excel cell value to float, or None if absent/non-numeric.
    Normalizes comma decimals and stray spaces/NBSP before parsing."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\xa0", "").replace(" ", "").replace(",", ".")
        if value == "":
            return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _to_int(value) -> int | None:
    val = _to_float(value)
    return None if val is None else int(val)


def _coerce_int_cell(raw):
    """Return int, None (truly blank/merged), or _UNPARSEABLE (present but
    not numeric) — the three-way distinction breakdown_inheriting needs."""
    if raw is None:
        return None
    val = _to_int(raw)
    return val if val is not None else _UNPARSEABLE


def _cell_int(row, col, default: int | None = None) -> int | None:
    """Read row[col] as int, falling back to default when absent/non-numeric."""
    val = _to_int(row[col]) if len(row) > col else None
    return val if val is not None else default


def _find_consecutive_run(
    ws, scan_start: int, scan_end: int, col_lo: int, col_hi: int, n: int
) -> dict[int, int]:
    """Return {col_idx: value} for the last row in [scan_start, scan_end]
    whose columns in [col_lo, col_hi) hold exactly the integers 1..n,
    consecutively and without duplicates. {} if no such row exists."""
    last_match: dict[int, int] = {}
    for row in ws.iter_rows(min_row=scan_start, max_row=scan_end, values_only=True):
        candidates = []
        for col_idx in range(col_lo, min(col_hi, len(row))):
            int_val = _to_int(row[col_idx])
            if int_val is not None and 1 <= int_val <= n:
                candidates.append((col_idx, int_val))
        if len(candidates) != n:
            continue
        candidates.sort(key=lambda x: x[1])
        values = [v for _, v in candidates]
        if len(set(values)) != n:
            continue
        if all(candidates[i + 1][1] == candidates[i][1] + 1 for i in range(n - 1)):
            last_match = dict(candidates)
    return last_match


def _find_group_header_col(ws, start_row: int, labels: tuple[str, ...]) -> int | None:
    scan_start = max(1, start_row - 10)
    for row in ws.iter_rows(min_row=scan_start, max_row=start_row):
        for cell in row:
            if isinstance(cell.value, str):
                norm = _normalize_text(cell.value)
                if any(label in norm for label in labels):
                    return cell.column - 1
    return None


def _label_semester_maps(
    ws, start_row: int, n: int
) -> tuple[dict[int, int], dict[int, int]]:
    """Primary strategy (map §4.3): find the weekly-hours and credit group
    headers, then read the local 1..N sub-number row beneath each.

    The N sub-columns aren't necessarily adjacent — one observed file spaces
    them 3 apart (with a "kurs" grouping column above) — so the search window
    for a group with no known next-group boundary is generous rather than a
    tight `n`-based guess.
    """
    scan_start = max(1, start_row - 10)
    weekly_col = _find_group_header_col(ws, start_row, WEEKLY_GROUP_LABELS)
    credit_col = _find_group_header_col(ws, start_row, CREDIT_GROUP_LABELS)
    generous_span = n * 6 + 5

    weekly_map: dict[int, int] = {}
    if weekly_col is not None:
        hi = credit_col if credit_col is not None else weekly_col + generous_span
        weekly_map = _find_consecutive_run(ws, scan_start, start_row, weekly_col, hi, n)

    credit_map: dict[int, int] = {}
    if credit_col is not None:
        credit_map = _find_consecutive_run(
            ws, scan_start, start_row, credit_col, credit_col + generous_span, n
        )
    return weekly_map, credit_map


def _detect_col_range(
    ws, start_row: int, lo: int, hi: int, offset: int
) -> dict[int, int]:
    """Fallback strategy (map §4.4): scan header rows near the anchor for the
    logical column-number row (1, 2, …, 29), not present in every file."""
    scan_start = max(1, start_row - 10)
    last_match = None
    for row in ws.iter_rows(min_row=scan_start, max_row=start_row, values_only=True):
        candidates = []
        for col_idx, cell_val in enumerate(row):
            int_val = _to_int(cell_val)
            if int_val is not None and lo <= int_val <= hi:
                candidates.append((col_idx, int_val))

        if len(candidates) < 3:
            continue
        values = [v for _, v in candidates]
        if len(values) != len(set(values)):
            continue
        candidates.sort(key=lambda x: x[1])
        is_consecutive = all(
            candidates[i + 1][1] == candidates[i][1] + 1
            for i in range(len(candidates) - 1)
        )
        if not is_consecutive:
            continue

        last_match = {col_idx: int_val - offset for col_idx, int_val in candidates}

    return last_match or {}


def _fallback_semester_maps(
    ws, start_row: int, n: int
) -> tuple[dict[int, int], dict[int, int]]:
    weekly = _detect_col_range(ws, start_row, 12, 11 + n, 11)
    credit = _detect_col_range(ws, start_row, 12 + n, 11 + 2 * n, 11 + n)
    return weekly, credit


def _validate_semester_maps(
    weekly: dict[int, int], credit: dict[int, int], n: int
) -> None:
    if len(weekly) != n or len(credit) != n:
        raise ParseError(
            "semestr ustunlari to'liq emas: "
            f"haftalik={sorted(weekly)}, kredit={sorted(credit)}, kutilgan soni={n}"
        )
    overlap = set(weekly) & set(credit)
    if overlap:
        raise ParseError(f"haftalik va kredit ustunlari kesishadi: {sorted(overlap)}")
    if max(weekly) >= min(credit):
        raise ParseError(
            "kredit ustunlari haftalik ustunlaridan o'ngda bo'lishi kerak: "
            f"haftalik={sorted(weekly)}, kredit={sorted(credit)}"
        )


def detect_semester_column_maps(
    ws, start_row: int, n: int
) -> tuple[dict[int, int], dict[int, int]]:
    """Return (weekly-hours map, credit map): {0-based col: semester number}.

    Tries the label-driven strategy first (map §4.3), falling back to the
    logical-column-number row (map §4.4, absent in some files). Raises
    ParseError when neither strategy nor the fallback yields a full,
    non-overlapping pair of maps.
    """
    weekly_map, credit_map = _label_semester_maps(ws, start_row, n)
    if not weekly_map or not credit_map:
        fb_weekly, fb_credit = _fallback_semester_maps(ws, start_row, n)
        weekly_map = weekly_map or fb_weekly
        credit_map = credit_map or fb_credit
    if not weekly_map or not credit_map:
        raise ParseError("semestr ustunlari (haftalik soat / kredit) aniqlanmadi")
    _validate_semester_maps(weekly_map, credit_map, n)
    return weekly_map, credit_map


def _find_code_name_cols(ws, num_col: int, start_row: int) -> tuple[int, int]:
    """Find code/name columns from the first course row below the anchor;
    fall back to num_col+1 / num_col+2 when the row can't be located."""
    code_col, name_col = None, None
    for data_row in ws.iter_rows(
        min_row=start_row + 1, max_row=start_row + 10, values_only=True
    ):
        if not data_row or data_row[num_col] is None:
            continue
        if not re.match(r"^\d+\.\d+", str(data_row[num_col]).strip().rstrip(".")):
            continue
        found = []
        for i in range(num_col + 1, len(data_row)):
            if data_row[i] is not None and len(found) < 2:
                found.append(i)
        if len(found) >= 2:
            code_col, name_col = found[0], found[1]
        elif len(found) == 1:
            code_col = name_col = found[0]
        break

    if code_col is None:
        code_col = num_col + 1
    if name_col is None:
        name_col = num_col + 2
    return code_col, name_col


def find_total_hours_col(ws, auditoriya_row: int, start_row: int) -> int:
    """Locate the total-hours column, searching only the header band right
    around the 'auditoriya' label so an earlier, unrelated 'soat' cell in the
    metadata or calendar zones can't hijack the column."""
    band_start = max(1, auditoriya_row - 2)
    col = find_header_col(ws, "umumiy yuklama", start_row, min_row=band_start)
    if col is None:
        col = find_header_col(ws, "soat", start_row, min_row=band_start)
    if col is None:
        raise ParseError("'Umumiy yuklama'/'soat' ustuni topilmadi")
    return col


def _find_classroom_col(ws, start_row: int) -> tuple[int, int]:
    """Find the classroom ('auditoriya') column and the header row it sits
    on, as (classroom_col, auditoriya_row)."""
    for hrow in ws.iter_rows(max_row=start_row):
        for cell in hrow:
            if isinstance(cell.value, str) and "auditoriya" in _normalize_text(
                cell.value
            ):
                return cell.column - 1, cell.row
    raise ParseError("'Auditoriya' ustuni topilmadi")


def _find_subcategory_cols(
    ws, auditoriya_row: int, classroom_col: int, warnings: list[str]
) -> tuple[int, int, int | None, int | None, int | None]:
    """Map the label row below 'auditoriya' to the five subcategory columns.

    'Ma'ruza' and 'Amaliy' are always present (map §4.3) — their absence is
    structural. 'Laboratoriya', 'Seminar', and 'Kurs ishi' are legitimately
    absent on some sheets; an absent one means 0 hours, not a guessed offset.
    """
    lecture_col = practice_col = lab_col = seminar_col = course_proj_col = None
    label_row = next(
        ws.iter_rows(
            min_row=auditoriya_row + 1, max_row=auditoriya_row + 1, values_only=True
        )
    )
    for i, v in enumerate(label_row):
        if i <= classroom_col or not isinstance(v, str):
            continue
        norm = re.sub(r"[^a-z]", "", _normalize_text(v))
        if "maruza" in norm and lecture_col is None:
            lecture_col = i
        elif "amaliy" in norm and practice_col is None:
            practice_col = i
        elif ("laborator" in norm or "labarator" in norm) and lab_col is None:
            lab_col = i
        elif "seminar" in norm and seminar_col is None:
            seminar_col = i
        elif "kurs" in norm and course_proj_col is None:
            course_proj_col = i

    if lecture_col is None or practice_col is None:
        raise ParseError("'Ma'ruza'/'Amaliy' ustunlari topilmadi")
    for label, col in (
        ("Laboratoriya", lab_col),
        ("Seminar", seminar_col),
        ("Kurs ishi", course_proj_col),
    ):
        if col is None:
            warnings.append(f"'{label}' ustuni topilmadi; 0 soat deb hisoblandi")
    return lecture_col, practice_col, lab_col, seminar_col, course_proj_col


def detect_course_columns(
    ws, num_col: int, start_row: int, warnings: list[str]
) -> ColumnLayout:
    """Return the course ColumnLayout by orchestrating the focused detectors."""
    code_col, name_col = _find_code_name_cols(ws, num_col, start_row)
    classroom_col, auditoriya_row = _find_classroom_col(ws, start_row)
    hours_col = find_total_hours_col(ws, auditoriya_row, start_row)
    lecture_col, practice_col, lab_col, seminar_col, course_proj_col = (
        _find_subcategory_cols(ws, auditoriya_row, classroom_col, warnings)
    )

    return ColumnLayout(
        num=num_col,
        code=code_col,
        name=name_col,
        hours=hours_col,
        classroom=classroom_col,
        lecture=lecture_col,
        practice=practice_col,
        lab=lab_col,
        seminar=seminar_col,
        course_proj=course_proj_col,
    )


def _round_half_up(value: float) -> int:
    return math.floor(value + 0.5)


def map_semester_values(
    row: tuple, semester_col_map: dict[int, int], label: str, warnings: list[str]
) -> dict[int, int]:
    """Return {semester_number: value} for non-zero semester cells.

    Shared by credit and weekly-hours extraction. A non-integral value (e.g.
    a 1.5-credit course) is warned about and rounded half-up rather than
    silently truncated.
    """
    result = {}
    for col_idx, sem_num in semester_col_map.items():
        raw = row[col_idx] if len(row) > col_idx else None
        val = _to_float(raw)
        if not val:
            continue
        rounded = _round_half_up(val)
        if rounded != val:
            warnings.append(
                f"{label}: {sem_num}-semestr qiymati {raw!r} {rounded} ga yaxlitlandi"
            )
        result[sem_num] = rounded
    return result


def _clean_name(value) -> str:
    """Normalize a course/alternative name read from a cell.

    Ministry templates use hard line breaks (Alt+Enter) in cells; a stray
    newline or '|' inside a name would terminate a Markdown table row and
    corrupt every row below it. Collapse all internal whitespace to single
    spaces and replace '|' so the name is always row-safe.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("|", " ")).strip()


def _is_recognized_skip_row(row) -> bool:
    return any(
        isinstance(cell, str)
        and any(label in _normalize_text(cell) for label in SKIP_ROW_LABELS)
        for cell in row
    )


def _iter_section_rows(ws, num_col: int, section: int, start_row: int):
    """Yield (num_str, row) for each row in the given top-level section (1 or 2).

    Enters when the section's marker is seen and stops at the first 'Jami'
    total row or the next section marker. This is the single shared
    row-walk/stop skeleton for both parse functions. Scanning starts at the
    '1.00' anchor row, never from row 1 — the metadata and academic-calendar
    zones above it can hold stray cells that normalize to a marker (e.g. a
    lone week number '1'), which would otherwise trigger a false section
    entry.
    """
    in_section = False
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        num = row[num_col] if len(row) > num_col else None
        num_str = str(num).strip().rstrip(".") if num is not None else None
        marker = _normalize_section_marker(num)

        if not in_section:
            if marker == str(section):
                in_section = True
            continue

        if any(
            cell is not None and str(cell).strip().rstrip(":").casefold() == "jami"
            for cell in row
        ):
            return
        if marker is not None:
            return

        yield num_str, row


def breakdown_from(row, cols: tuple, fill: int) -> dict:
    """Read the six hour-breakdown fields, defaulting every missing cell to
    `fill`. A None column (no header for that hour type) always reads as 0."""
    return {
        f: (0 if c is None else _cell_int(row, c, fill))
        for f, c in zip(BREAKDOWN_FIELDS, cols)
    }


def breakdown_inheriting(
    row, cols: tuple, fallbacks: dict, label: str, warnings: list[str]
) -> dict:
    """Read the six hour-breakdown fields, inheriting per-field fallbacks.

    Used for a selective continuation row, whose numeric cells are vertically
    merged with the slot's first row: a truly blank (None) cell inherits the
    slot's value. A present-but-unparseable cell does NOT inherit — that
    would silently pass off a bad cell as the slot's real value — it reads as
    0 and is warned about instead.
    """
    result = {}
    for f, c in zip(BREAKDOWN_FIELDS, cols):
        if c is None:
            result[f] = 0
            continue
        raw = row[c] if len(row) > c else None
        coerced = _coerce_int_cell(raw)
        if coerced is None:
            result[f] = fallbacks[f]
        elif coerced is _UNPARSEABLE:
            warnings.append(
                f"{label}: {f} katakchasi o'qilmadi ({raw!r}); 0 deb hisoblandi"
            )
            result[f] = 0
        else:
            result[f] = coerced
    return result


def _append_alternative(
    slot: SelectiveSlot, code_str: str, name_str: str, breakdown: dict
) -> None:
    """Append an Alternative to the slot, skipping fully blank continuation rows."""
    if code_str or name_str:
        slot.alternatives.append(Alternative(code=code_str, name=name_str, **breakdown))


def parse_selective_courses(
    ws,
    layout: ColumnLayout,
    semester_map: dict[int, int],
    weekly_map: dict[int, int],
    start_row: int,
    warnings: list[str],
) -> list[SelectiveSlot]:
    """Parse section 2 (tanlov fanlar / selective courses) into SelectiveSlots.

    Each slot carries the shared hours/credits plus an `alternatives` list.
    Numeric columns are vertically merged in the spreadsheet, so their values
    appear only on the slot's first row; continuation rows carry only code and
    name and inherit the slot's breakdown for any merged (None) cell.
    """
    slots: list[SelectiveSlot] = []
    current: SelectiveSlot | None = None
    slot_breakdown: dict = {}

    for num_str, row in _iter_section_rows(ws, layout.num, 2, start_row):
        code = row[layout.code] if len(row) > layout.code else None
        name = row[layout.name] if len(row) > layout.name else None
        code_str = str(code).strip() if code is not None else ""
        name_str = _clean_name(name)

        if is_course_number(num_str) and section_prefix(num_str) == "2":
            hours_raw = row[layout.hours] if len(row) > layout.hours else None
            current = SelectiveSlot(
                num=num_str,
                hours=_to_int(hours_raw),
                semester_credits=map_semester_values(
                    row, semester_map, num_str, warnings
                ),
                semester_weekly_hours=map_semester_values(
                    row, weekly_map, num_str, warnings
                ),
            )
            slots.append(current)
            # The slot row's breakdown is the default every continuation row
            # inherits where its cells are merged (None).
            slot_breakdown = breakdown_from(row, layout.breakdown, 0)
            _append_alternative(current, code_str, name_str, slot_breakdown)

        elif current is not None and num_str is None:
            _append_alternative(
                current,
                code_str,
                name_str,
                breakdown_inheriting(
                    row, layout.breakdown, slot_breakdown, current.num, warnings
                ),
            )

        elif not _is_recognized_skip_row(row):
            warnings.append(f"tanilmagan qator o'tkazib yuborildi: {row[:3]!r}")

    if not slots:
        warnings.append("tanlov fanlar topilmadi")

    return slots


def parse_core_courses(
    ws,
    layout: ColumnLayout,
    semester_map: dict[int, int],
    weekly_map: dict[int, int],
    start_row: int,
    warnings: list[str],
) -> list[Course]:
    """Parse section 1 (majburiy fanlar / core courses) into Courses."""
    courses: list[Course] = []

    for num_str, row in _iter_section_rows(ws, layout.num, 1, start_row):
        if not (is_course_number(num_str) and section_prefix(num_str) == "1"):
            if not _is_recognized_skip_row(row):
                warnings.append(f"tanilmagan qator o'tkazib yuborildi: {row[:3]!r}")
            continue
        code = row[layout.code] if len(row) > layout.code else None
        name = row[layout.name] if len(row) > layout.name else None
        hours_raw = row[layout.hours] if len(row) > layout.hours else None
        course = Course(
            num=num_str,
            code=str(code).strip() if code else "",
            name=_clean_name(name),
            hours=_to_int(hours_raw),
            **breakdown_from(row, layout.breakdown, 0),
            semester_credits=map_semester_values(row, semester_map, num_str, warnings),
            semester_weekly_hours=map_semester_values(
                row, weekly_map, num_str, warnings
            ),
        )
        courses.append(course)

    return courses


def _section_marker_hours(
    ws, num_col: int, hours_col: int, section: int, start_row: int
) -> int | None:
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        num = row[num_col] if len(row) > num_col else None
        if _normalize_section_marker(num) == str(section):
            return _to_int(row[hours_col]) if len(row) > hours_col else None
    return None


def _validate_section_totals(
    ws,
    layout: ColumnLayout,
    start_row: int,
    core: list[Course],
    slots: list[SelectiveSlot],
    warnings: list[str],
) -> None:
    """Cross-check the marker rows' own aggregate against the parsed courses
    (map §5.1) — the strongest available signal that the hours columns were
    detected correctly."""
    core_total = sum(c.hours for c in core if c.hours is not None)
    marker1 = _section_marker_hours(ws, layout.num, layout.hours, 1, start_row)
    if marker1 is not None and marker1 != core_total:
        warnings.append(
            f"1-bo'lim jami soat mos kelmadi: varaqda {marker1},"
            f" hisoblangan {core_total}"
        )

    slot_total = sum(s.hours for s in slots if s.hours is not None)
    marker2 = _section_marker_hours(ws, layout.num, layout.hours, 2, start_row)
    if marker2 is not None and marker2 != slot_total:
        warnings.append(
            f"2-bo'lim jami soat mos kelmadi: varaqda {marker2},"
            f" hisoblangan {slot_total}"
        )


def _check_classroom_breakdown(
    label: str,
    classroom: int,
    lecture: int,
    practice: int,
    lab: int,
    seminar: int,
    warnings: list[str],
) -> None:
    total = lecture + practice + lab + seminar
    if classroom != total:
        warnings.append(
            f"{label}: auditoriya {classroom} != ma'ruza+amaliy+lab+seminar ({total})"
        )


def _validate_classroom_breakdown(
    core: list[Course], slots: list[SelectiveSlot], warnings: list[str]
) -> None:
    for course in core:
        _check_classroom_breakdown(
            f"{course.num} {course.name}",
            course.classroom,
            course.lecture,
            course.practice,
            course.lab,
            course.seminar,
            warnings,
        )
    for slot in slots:
        for alt in slot.alternatives:
            _check_classroom_breakdown(
                f"{slot.num} {alt.name}",
                alt.classroom,
                alt.lecture,
                alt.practice,
                alt.lab,
                alt.seminar,
                warnings,
            )


def _check_credit_hours(
    label: str, hours: int | None, semester_credits: dict[int, int], warnings: list[str]
) -> None:
    if hours is None:
        return
    if not semester_credits:
        if hours:
            warnings.append(f"{label}: soat {hours} bor, lekin semestr kreditlari yo'q")
        return
    remainder = hours % 30
    if remainder != 0:
        warnings.append(
            f"{label}: soat {hours} 30 ga qoldiqsiz bo'linmaydi (qoldiq {remainder})"
        )
    expected = hours // 30
    actual = sum(semester_credits.values())
    if expected != actual:
        warnings.append(
            f"{label}: kredit {expected} (soat/30) semestrlar yig'indisi {actual}"
            " bilan mos kelmadi"
        )


def _validate_credit_hours(
    core: list[Course], slots: list[SelectiveSlot], warnings: list[str]
) -> None:
    for course in core:
        _check_credit_hours(
            f"{course.num} {course.name}",
            course.hours,
            course.semester_credits,
            warnings,
        )
    for slot in slots:
        name = slot.alternatives[0].name if slot.alternatives else ""
        _check_credit_hours(
            f"{slot.num} {name}", slot.hours, slot.semester_credits, warnings
        )


def _hours_column_all_blank(ws, layout: ColumnLayout, start_row: int) -> bool:
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        num = row[layout.num] if len(row) > layout.num else None
        if is_course_number(str(num).strip().rstrip(".") if num is not None else None):
            if len(row) > layout.hours and row[layout.hours] is not None:
                return False
    return True


def _check_uncalculated_formulas(
    path: Path, layout: ColumnLayout, start_row: int, warnings: list[str]
) -> None:
    """A workbook never recalculated by Excel caches formula cells as None
    under data_only=True — reopen without it to tell "genuinely blank" from
    "formula Excel never evaluated"."""
    wb_raw = openpyxl.load_workbook(path, data_only=False)
    ws_raw = wb_raw.worksheets[0]
    has_formula = any(
        len(row) > layout.hours
        and isinstance(row[layout.hours], str)
        and row[layout.hours].startswith("=")
        for row in ws_raw.iter_rows(min_row=start_row, values_only=True)
    )
    if has_formula:
        raise ParseError(
            "soat ustunidagi formulalar hisoblanmagan;"
            " faylni Excelda ochib qayta saqlang"
        )
    warnings.append("soat ustuni bo'sh: birorta kurs uchun ham soat topilmadi")


def parse_workbook(path: Path) -> ParseResult:
    """Parse one o'quv reja .xlsx into a ParseResult. The single entry point
    both `parser.main` and `plans.importer` build on."""
    warnings: list[str] = []
    wb = openpyxl.load_workbook(path, data_only=True)
    if len(wb.sheetnames) > 1:
        warnings.append(
            "varaqda bir nechta varaq bor; birinchisi "
            f"{wb.sheetnames[0]!r} o'qildi, qolganlari e'tiborsiz qoldirildi: "
            f"{wb.sheetnames[1:]}"
        )
    ws = wb.worksheets[0]

    direction = resolve_direction(ws, warnings)
    direction_code, direction_name = direction if direction is not None else ("", "")

    year_cell = find_cell_containing(ws, "quv yili")
    start_year = extract_start_year(str(year_cell.value)) if year_cell else ""
    if not start_year:
        warnings.append("o'quv yili topilmadi")

    degree, _duration_text, edu_type = read_metadata(ws, warnings)
    years = extract_program_years(ws)
    n = round(years * 2)

    num_col, start_row = _find_anchor(ws, warnings)
    layout = detect_course_columns(ws, num_col, start_row, warnings)
    weekly_map, credit_map = detect_semester_column_maps(ws, start_row, n)

    if _hours_column_all_blank(ws, layout, start_row):
        _check_uncalculated_formulas(path, layout, start_row, warnings)

    core = parse_core_courses(ws, layout, credit_map, weekly_map, start_row, warnings)
    slots = parse_selective_courses(
        ws, layout, credit_map, weekly_map, start_row, warnings
    )

    _validate_section_totals(ws, layout, start_row, core, slots, warnings)
    _validate_classroom_breakdown(core, slots, warnings)
    _validate_credit_hours(core, slots, warnings)

    return ParseResult(
        direction_code=direction_code,
        direction_name=direction_name,
        start_year=start_year,
        degree=degree,
        duration_years=years,
        edu_type=edu_type,
        core=core,
        slots=slots,
        warnings=warnings,
    )
