"""Tests for parser/parser.py, built on small in-memory openpyxl workbooks
that mirror the real layouts documented in docs/oquv_reja_map.md."""

import tempfile
from pathlib import Path

import openpyxl
from django.test import SimpleTestCase

from parser.models import ParseError
from parser.parser import parse_workbook


def _write_workbook(
    rows: dict[tuple[int, int], object], sheet_names: list[str] | None = None
) -> Path:
    """Build a workbook from {(row, col): value} (1-based) and save it to a
    temp file, returning its path. Extra sheets, if named, are left empty."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for (row, col), value in rows.items():
        ws.cell(row=row, column=col, value=value)
    for name in (sheet_names or [])[1:]:
        wb.create_sheet(name)
    if sheet_names:
        ws.title = sheet_names[0]
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return Path(tmp.name)


def _standard_layout(
    n: int, direction: str, duration: str, year: str
) -> dict[tuple[int, int], object]:
    """A minimal but complete sheet: metadata block, one core course, one
    selective slot with two alternatives, `n` semesters."""
    rows: dict[tuple[int, int], object] = {
        (2, 1): direction,
        (3, 1): f"O'qish muddati - {duration}",
        (4, 1): "Akademik daraja - BAKALAVR",
        (4, 2): "O'qish shakli - Kredit-modul",
        (5, 1): "Ta'lim shakli - kunduzgi",
        (6, 1): f"{year} o'quv yili",
        (17, 1): "II. O'QUV REJASI",
        (18, 4): "Umumiy yuklama",
        (18, 5): "Auditoriya",
        (18, 11): "Semestrdagi haftalar soni",
        (18, 19): "Semestrdagi kreditlar taqsimoti",
        (19, 6): "Ma'ruza",
        (19, 7): "Amaliy",
        (19, 8): "Laboratoriya",
        (19, 9): "Seminar",
        (19, 10): "Kurs ishi",
        (20, 1): "1.00",
        (20, 3): "Majburiy fanlar",
        (21, 1): "1.01",
        (21, 2): "COD1",
        (21, 3): "Course A",
        (21, 4): 120,
        (21, 5): 60,
        (21, 6): 30,
        (21, 7): 30,
        (22, 1): "2.00",
        (22, 3): "Tanlov fanlar",
        (23, 1): "2.01",
        (23, 2): "SEL1",
        (23, 3): "Selective A",
        (23, 4): 90,
        (23, 5): 45,
        (23, 6): 45,
        (24, 2): "SEL2",
        (24, 3): "Selective B",
        (25, 3): "Jami:",
    }
    for i in range(n):
        rows[(19, 11 + i)] = i + 1
        rows[(19, 19 + i)] = i + 1
    rows[(21, 11)] = 4  # course: semester 1 weekly hours
    rows[(21, 19)] = 4  # course: semester 1 credits
    rows[(23, 13)] = 3  # selective slot: semester 3 weekly hours
    rows[(23, 21)] = 3  # selective slot: semester 3 credits
    return rows


class ParseWorkbookHappyPathTest(SimpleTestCase):
    def test_toliq_varaqni_ochadi(self) -> None:
        path = _write_workbook(
            _standard_layout(
                8,
                "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim",
                "4 yil",
                "2024-2025",
            )
        )
        result = parse_workbook(path)
        self.assertEqual(result.direction_code, "60110500")
        self.assertEqual(result.direction_name, "Boshlang'ich ta'lim")
        self.assertEqual(result.start_year, "2024")
        self.assertEqual(result.duration_years, 4)
        self.assertEqual(result.edu_type, "Kunduzgi")
        self.assertEqual(len(result.core), 1)
        self.assertEqual(result.core[0].semester_credits, {1: 4})
        self.assertEqual(len(result.slots), 1)
        self.assertEqual(len(result.slots[0].alternatives), 2)

    def test_slash_yil_ham_ishlaydi(self) -> None:
        path = _write_workbook(
            _standard_layout(
                8,
                "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim",
                "4 yil",
                "2023/2024",
            )
        )
        result = parse_workbook(path)
        self.assertEqual(result.start_year, "2023")

    def test_en_dash_yonalish(self) -> None:
        path = _write_workbook(
            _standard_layout(
                8,
                "Ta'lim yo'nalishi: 60110500 – Boshlang'ich ta'lim",
                "4 yil",
                "2024-2025",
            )
        )
        result = parse_workbook(path)
        self.assertEqual(result.direction_code, "60110500")
        self.assertEqual(result.direction_name, "Boshlang'ich ta'lim")

    def test_uch_yillik_dastur_oltita_semestr(self) -> None:
        path = _write_workbook(
            _standard_layout(
                6,
                "Ta'lim yo'nalishi: 60111200 - Jismoniy madaniyat",
                "3 yil",
                "2024-2025",
            )
        )
        result = parse_workbook(path)
        self.assertEqual(result.duration_years, 3)
        self.assertEqual(result.core[0].semester_credits, {1: 4})

    def test_faol_varaq_birinchisi_bolmasa_ham_birinchisi_oqiladi(self) -> None:
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        path = _write_workbook(rows, sheet_names=["Kunduzgi", "Sirtqi"])
        wb = openpyxl.load_workbook(path)
        wb.active = 1
        wb.save(path)
        result = parse_workbook(path)
        self.assertEqual(result.direction_code, "60110500")
        self.assertTrue(any("bir nechta varaq" in w for w in result.warnings))

    def test_vergulli_kredit_yaxlitlanadi_va_ogohlantiradi(self) -> None:
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        rows[(21, 19)] = "1,5"
        path = _write_workbook(rows)
        result = parse_workbook(path)
        self.assertEqual(result.core[0].semester_credits, {1: 2})
        self.assertTrue(any("1,5" in w for w in result.warnings))

    def test_talim_shakli_oqish_shakli_bilan_adashmaydi(self) -> None:
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        path = _write_workbook(rows)
        result = parse_workbook(path)
        self.assertEqual(result.edu_type, "Kunduzgi")

    def test_talim_shakli_apostrof_variantlari(self) -> None:
        for label in ("Ta’lim shakli - kunduzgi", "Ta‘lim shakli - kunduzgi"):
            rows = _standard_layout(
                8,
                "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim",
                "4 yil",
                "2024-2025",
            )
            rows[(5, 1)] = label
            path = _write_workbook(rows)
            result = parse_workbook(path)
            self.assertEqual(result.edu_type, "Kunduzgi")

    def test_notanish_talim_shakli_ogohlantiradi(self) -> None:
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        rows[(5, 1)] = "Ta'lim shakli - masofaviy"
        path = _write_workbook(rows)
        result = parse_workbook(path)
        self.assertEqual(result.edu_type, "")
        self.assertTrue(any("ta'lim shakli tanilmadi" in w for w in result.warnings))

    def test_kurs_loyihasi_ishi_yorlig_topiladi(self) -> None:
        rows = _standard_layout(
            8,
            "Ta'lim yo'nalishi: 60610100 - Axborot texnologiyalari",
            "4 yil",
            "2024-2025",
        )
        rows[(19, 10)] = "Kurs loyihasi (ishi)"
        rows[(21, 10)] = 20
        path = _write_workbook(rows)
        result = parse_workbook(path)
        self.assertEqual(result.core[0].course_proj, 20)

    def test_ustun_raqamlari_qatori_yoq_bolsada_yorliq_orqali_topiladi(self) -> None:
        """Map §4.4: the logical column-number row isn't present in every
        file — detection must not depend on it."""
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        path = _write_workbook(rows)
        result = parse_workbook(path)
        self.assertEqual(len(result.core), 1)

    def test_fallback_ustun_raqami_qatoridan_topiladi(self) -> None:
        """When the group-header labels are absent, fall back to the
        logical column-number row (map §4.4)."""
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        del rows[(18, 11)]
        del rows[(18, 19)]
        # Logical column numbers: weekly = 12..19, credit = 20..27 (offset 11).
        for i in range(8):
            rows[(19, 11 + i)] = 12 + i
            rows[(19, 19 + i)] = 20 + i
        path = _write_workbook(rows)
        result = parse_workbook(path)
        self.assertEqual(len(result.core), 1)
        self.assertEqual(result.core[0].semester_credits, {1: 4})


class ParseWorkbookFailurePathTest(SimpleTestCase):
    def test_1_00_belgisi_yoq_xato(self) -> None:
        # No course table at all: nothing in the sheet normalizes to a
        # section marker, so no anchor candidate exists.
        rows = {
            (2, 1): "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim",
            (3, 1): "O'qish muddati - 4 yil",
            (6, 1): "2024-2025 o'quv yili",
            (7, 1): "II. O'QUV REJASI",
        }
        path = _write_workbook(rows)
        with self.assertRaises(ParseError):
            parse_workbook(path)

    def test_oqish_muddati_yoq_xato(self) -> None:
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        del rows[(3, 1)]
        path = _write_workbook(rows)
        with self.assertRaises(ParseError):
            parse_workbook(path)

    def test_semestr_ustunlari_topilmasa_xato(self) -> None:
        rows = _standard_layout(
            8, "Ta'lim yo'nalishi: 60110500 - Boshlang'ich ta'lim", "4 yil", "2024-2025"
        )
        for key in list(rows):
            row, col = key
            if row in (18, 19) and col >= 11:
                del rows[key]
        path = _write_workbook(rows)
        with self.assertRaises(ParseError):
            parse_workbook(path)
